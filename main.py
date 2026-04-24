from fastapi import FastAPI, Request, Depends, HTTPException, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, case
from starlette.middleware.sessions import SessionMiddleware
from passlib.context import CryptContext
from datetime import date
import os, io, re
from dotenv import load_dotenv

from database import get_db, init_db, User, Content, PriceTable, Document, StudioRental, CustomerContact
import docgen

load_dotenv()

app = FastAPI(title="KICPA 콘텐츠 관리")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SECRET_KEY", "kicpa-dev-secret"))

@app.exception_handler(Exception)
async def all_exception_handler(request: Request, exc: Exception):
    import traceback
    from fastapi.responses import HTMLResponse
    detail = traceback.format_exc()
    return HTMLResponse(
        content=f"<h2 style='color:red'>오류 발생</h2><pre style='background:#f8f8f8;padding:20px'>{detail}</pre>",
        status_code=500
    )
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
pwd_ctx = CryptContext(schemes=["bcrypt"])

MONTHS = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
MONTH_ORDER = {m: i for i, m in enumerate(MONTHS)}

# ── 헬퍼 ────────────────────────────────────────
def get_user(request: Request): return request.session.get("user")
def require_login(request: Request):
    if not get_user(request): raise HTTPException(status_code=302, headers={"Location": "/login"})
def require_admin(request: Request):
    u = get_user(request)
    if not u or u["role"] != "admin": raise HTTPException(status_code=403, detail="권한 없음")

def fmt_date(d): return d.strftime("%Y.%m.%d") if d else ""
def clean_name(name):
    if not name: return ''
    for line in str(name).split('\n'):
        cleaned = re.sub(r'\[.*?\]', '', line).strip()
        if cleaned: return cleaned
    return re.sub(r'\[.*?\]', '', str(name).replace('\n', ' ')).strip()

def normalize_month(val):
    """다양한 월 표기를 'X월' 형식으로 통일
    예: '4' → '4월', '04' → '4월', '4월 ' → '4월', '4월청구' → '4월', '5월 청구' → '5월'
    """
    if not val: return None
    s = str(val).strip()
    # 'X월' 패턴 먼저 탐색 (예: '5월 청구', '5월청구', '12월')
    m = re.search(r'(\d{1,2})월', s)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return f"{num}월"
    # 순수 숫자만 있는 경우 (예: '5', '05')
    m = re.match(r'^(\d{1,2})$', s)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return f"{num}월"
    return s  # 매칭 안 되면 원본 반환

templates.env.filters["fmt_date"] = fmt_date
templates.env.filters["clean_name"] = clean_name
templates.env.filters["normalize_month"] = normalize_month
templates.env.globals["MONTHS"] = MONTHS
templates.env.globals["normalize_month"] = normalize_month

@app.on_event("startup")
def startup(): init_db()

# ── 인증 ────────────────────────────────────────
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if get_user(request): return RedirectResponse("/", 302)
    return templates.TemplateResponse("login.html", {"request": request, "error": ""})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(User).filter_by(username=username, is_active=True).first()
    if user and pwd_ctx.verify(password, user.password):
        request.session["user"] = {"id": user.id, "name": user.name, "role": user.role, "username": user.username}
        return RedirectResponse("/", 302)
    return templates.TemplateResponse("login.html", {"request": request, "error": "아이디 또는 비밀번호가 올바르지 않습니다."})

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", 302)

# ── 대시보드 ─────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, year: int = 2026, db: Session = Depends(get_db)):
    require_login(request)
    base = db.query(Content).filter_by(year=year)
    total   = base.count()
    shot    = base.filter(Content.shooting_date != None).count()
    opened  = base.filter(Content.open_date != None).count()
    billed  = base.filter(Content.billing != None, Content.billing != "").count()

    depts_raw = db.query(Content.department).filter_by(year=year).filter(Content.department != None).distinct().all()
    dept_list = [d[0] for d in depts_raw]
    dept_summary = []
    for dept in dept_list:
        q = db.query(Content).filter_by(year=year, department=dept)
        dept_summary.append({
            "department": dept,
            "total": q.count(),
            "shot": q.filter(Content.shooting_date != None).count(),
            "opened": q.filter(Content.open_date != None).count(),
        })
    dept_summary.sort(key=lambda x: x["total"], reverse=True)

    recent = db.query(Content).filter_by(year=year).order_by(Content.id.desc()).limit(8).all()
    return templates.TemplateResponse("dashboard.html", {
        "request": request, "user": get_user(request),
        "year": year, "total": total, "shot": shot, "opened": opened, "billed": billed,
        "dept_summary": dept_summary, "recent": recent,
    })

# ── 콘텐츠 목록 ───────────────────────────────────
@app.get("/contents", response_class=HTMLResponse)
def contents(request: Request, year: int = 2026, dept: str = "", month: str = "",
             fmt: str = "", billing: str = "", search: str = "",
             page: int = 1, db: Session = Depends(get_db)):
    require_login(request)
    per = 20
    q = db.query(Content).filter_by(year=year)
    if dept:    q = q.filter_by(department=dept)
    if month:   q = q.filter_by(shooting_month=month)
    if fmt:     q = q.filter(Content.shooting_format.ilike(f"%{fmt}%"))
    if billing == "Y": q = q.filter(Content.billing != None, Content.billing != "")
    if billing == "N": q = q.filter(or_(Content.billing == None, Content.billing == ""))
    if search:  q = q.filter(Content.course_name.ilike(f"%{search}%"))
    total = q.count()
    # 청구월 우선 정렬 → 촬영날짜(없으면 동영상 마킹) 순
    from sqlalchemy import text as sa_text
    billing_month_sort = case(
        {m: i for i, m in enumerate(MONTHS)},
        value=Content.billing_month, else_=99)
    date_sort = sa_text(
        "COALESCE(shooting_date, "
        "  CASE WHEN video_marking ~ '^[0-9]{4}[./-][0-9]' "
        "  THEN TO_DATE(SUBSTRING(REPLACE(REPLACE(video_marking,'.', '-'),'/','-'),1,10),'YYYY-MM-DD') "
        "  ELSE NULL END"
        ") DESC NULLS LAST"
    )
    rows  = q.order_by(billing_month_sort.desc(), date_sort, Content.id.desc()).offset((page-1)*per).limit(per).all()
    total_pages = (total + per - 1) // per

    depts   = [d[0] for d in db.query(Content.department).filter_by(year=year).filter(Content.department != None).distinct().order_by(Content.department).all()]
    formats = [f[0] for f in db.query(Content.shooting_format).filter_by(year=year).filter(Content.shooting_format != None).distinct().all()]
    return templates.TemplateResponse("contents.html", {
        "request": request, "user": get_user(request),
        "year": year, "rows": rows, "total": total, "page": page, "total_pages": total_pages,
        "depts": depts, "formats": formats,
        "dept": dept, "month": month, "fmt": fmt, "billing": billing, "search": search,
    })

# ── 콘텐츠 등록/수정 ──────────────────────────────
@app.get("/content/edit", response_class=HTMLResponse)
def content_edit_page(request: Request, id: int = 0,
                      f_page: int = 1, f_dept: str = "", f_month: str = "",
                      f_fmt: str = "", f_billing: str = "", f_search: str = "",
                      db: Session = Depends(get_db)):
    require_login(request)
    row = db.query(Content).filter_by(id=id).first() if id else None
    return templates.TemplateResponse("content_edit.html", {
        "request": request, "user": get_user(request), "row": row, "MONTHS": MONTHS, "msg": "",
        "f_page": f_page, "f_dept": f_dept, "f_month": f_month,
        "f_fmt": f_fmt, "f_billing": f_billing, "f_search": f_search,
    })

@app.post("/content/edit")
def content_edit_save(request: Request, id: int = Form(0), year: int = Form(2026),
    f_page: int = Form(1), f_dept: str = Form(""), f_month: str = Form(""),
    f_fmt: str = Form(""), f_billing: str = Form(""), f_search: str = Form(""),
    shooting_month: str=Form(""), course_name: str=Form(""), required_optional: str=Form(""),
    original_code: str=Form(""), category: str=Form(""), course_code: str=Form(""),
    session_count: str=Form(""), chapter_count: str=Form(""),
    instructor: str=Form(""), department: str=Form(""), kicpa_manager: str=Form(""),
    filming_consent: str=Form(""), shooting_date: str=Form(""), shooting_time: str=Form(""),
    shooting_format: str=Form(""), location: str=Form(""), has_quiz: str=Form(""),
    quiz_count: str=Form(""), materials_supply: str=Form(""), video_marking: str=Form(""),
    dev_outsource_date: str=Form(""), inspection_date: str=Form(""), open_date: str=Form(""),
    billing: str=Form(""), billing_month: str=Form(""), custom_price: str=Form(""),
    travel_hours: str=Form(""), travel_days: str=Form(""), travel_expense: str=Form(""),
    notes: str=Form(""), db: Session = Depends(get_db)):
    require_login(request)

    def to_date(s):
        try: return date.fromisoformat(s) if s else None
        except: return None
    def to_int(s):
        try: return int(s) if s else None
        except: return None

    data = dict(shooting_month=shooting_month or None, course_name=course_name or None,
        required_optional=required_optional or None, original_code=original_code or None,
        category=category or None, course_code=course_code or None,
        session_count=to_int(session_count), chapter_count=to_int(chapter_count),
        instructor=instructor or None, department=department or None,
        kicpa_manager=kicpa_manager or None, filming_consent=filming_consent or None,
        shooting_date=to_date(shooting_date), shooting_time=shooting_time or None,
        shooting_format=shooting_format or None, location=location or None,
        has_quiz=has_quiz or None, quiz_count=to_int(quiz_count),
        materials_supply=materials_supply or None, video_marking=video_marking or None,
        dev_outsource_date=to_date(dev_outsource_date), inspection_date=to_date(inspection_date),
        open_date=to_date(open_date), billing=billing or None,
        billing_month=normalize_month(billing_month),
        custom_price=to_int(custom_price),
        travel_hours=to_int(travel_hours),
        travel_days=to_int(travel_days),
        travel_expense=to_int(travel_expense), notes=notes or None)

    if id:
        db.query(Content).filter_by(id=id).update(data)
    else:
        db.add(Content(year=year, **data))
    db.commit()
    from urllib.parse import urlencode
    params = {"year": year, "page": f_page}
    if f_dept:    params["dept"]    = f_dept
    if f_month:   params["month"]   = f_month
    if f_fmt:     params["fmt"]     = f_fmt
    if f_billing: params["billing"] = f_billing
    if f_search:  params["search"]  = f_search
    return RedirectResponse("/contents?" + urlencode(params), 302)

# ── 촬영 일정 ─────────────────────────────────────
@app.get("/schedule", response_class=HTMLResponse)
def schedule(request: Request, year: int = 2026, month: str = "", dept: str = "", db: Session = Depends(get_db)):
    require_login(request)
    q = db.query(Content).filter_by(year=year).filter(Content.shooting_date != None)
    if month: q = q.filter_by(shooting_month=month)
    if dept:  q = q.filter_by(department=dept)
    rows = q.order_by(Content.shooting_date.desc()).all()
    grouped = {}
    for r in rows:
        key = r.shooting_date.isoformat()
        grouped.setdefault(key, []).append(r)
    depts = [d[0] for d in db.query(Content.department).filter_by(year=year).filter(Content.department != None).distinct().order_by(Content.department).all()]
    return templates.TemplateResponse("schedule.html", {
        "request": request, "user": get_user(request),
        "year": year, "grouped": grouped, "month": month, "dept": dept, "depts": depts,
    })

# ── Excel 내보내기 ─────────────────────────────────
@app.get("/export")
def export(request: Request, year: int = 2026, dept: str = "", month: str = "",
           billing: str = "", search: str = "", db: Session = Depends(get_db)):
    require_login(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(Content).filter_by(year=year)
    if dept:    q = q.filter_by(department=dept)
    if month:   q = q.filter_by(shooting_month=month)
    if billing == "Y": q = q.filter(Content.billing != None, Content.billing != "")
    if billing == "N": q = q.filter(or_(Content.billing == None, Content.billing == ""))
    if search:  q = q.filter(Content.course_name.ilike(f"%{search}%"))
    rows = q.order_by(Content.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "개발관리"
    ws.merge_cells("A1:Z1")
    ws["A1"] = f"{year}년 KICPA 콘텐츠개발 및 동영상 촬영 현황"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["No","촬영월","과정명","필수/선택","원코드","카테고리","과정코드",
               "차시수","챕터수","강사","담당부서","한공회담당","촬영동의서",
               "촬영날짜","촬영시간","촬영형식","장소","퀴즈유무","퀴즈문항수","교안수급",
               "동영상마킹","개발(외주)","검수","오픈일","비용청구","비고"]
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=i+1, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2F5496")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, r in enumerate(rows):
        vals = [ri+1, r.shooting_month, r.course_name, r.required_optional, r.original_code,
                r.category, r.course_code, r.session_count, r.chapter_count, r.instructor,
                r.department, r.kicpa_manager, r.filming_consent,
                r.shooting_date, r.shooting_time, r.shooting_format, r.location,
                r.has_quiz, r.quiz_count, r.materials_supply, r.video_marking,
                r.dev_outsource_date, r.inspection_date, r.open_date, r.billing, r.notes]
        for ci, v in enumerate(vals):
            ws.cell(row=ri+3, column=ci+1, value=v)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"{year}_콘텐츠개발_개발현황.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# ── Excel 가져오기 ─────────────────────────────────
@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    require_login(request)
    return templates.TemplateResponse("import.html", {"request": request, "user": get_user(request), "msg": "", "msg_type": ""})

@app.post("/import")
async def import_excel(request: Request, year: int = Form(2026), import_mode: str = Form("append"),
                       excel_file: UploadFile = File(...), db: Session = Depends(get_db)):
    require_login(request)
    import openpyxl
    from openpyxl.utils.datetime import from_excel

    content = await excel_file.read()
    buf = io.BytesIO(content)
    try:
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb["개발관리"] if "개발관리" in wb.sheetnames else wb.active

        # replace 모드: 삭제 전 수동 입력 데이터 저장 (과정명 기준 복원용)
        manual_data = {}
        if import_mode == "replace":
            for r in db.query(Content).filter_by(year=year).all():
                key = clean_name(r.course_name or "").strip()
                if key:
                    manual_data[key] = {
                        "custom_price":   r.custom_price,
                        "travel_hours":   r.travel_hours,
                        "travel_expense": r.travel_expense,
                        "notes":          r.notes,
                    }
            db.query(Content).filter_by(year=year).delete()
            db.commit()

        def to_date(v):
            if v is None: return None
            if isinstance(v, (date,)): return v
            try:
                import datetime
                if isinstance(v, datetime.datetime): return v.date()
                if isinstance(v, str): return date.fromisoformat(v[:10])
                if isinstance(v, (int, float)): return from_excel(v).date()
            except: pass
            return None

        prev_month = ""
        imported = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_name = str(row[2] or "").strip()
            if not raw_name: continue

            month_val = str(row[1] or "").strip()
            if month_val: prev_month = month_val

            # 과정명 정리: ==> 기준으로 분리 후 대괄호 제거
            name_parts = re.split(r'\s*\n*\s*==>', raw_name)
            name_cleaned = ""
            for line in name_parts[0].split('\n'):
                cleaned = re.sub(r'\[.*?\]', '', line).strip()
                if cleaned:
                    name_cleaned = cleaned
                    break
            if not name_cleaned:
                name_cleaned = re.sub(r'\[.*?\]', '', name_parts[0].replace('\n', ' ')).strip()

            # 비고: ==> 이후 내용 + 기존 엑셀 비고 병합
            extra_note = ("==>" + name_parts[1].strip()) if len(name_parts) > 1 else ""
            excel_note = str(row[26] or "").strip()   # 비고: 비용청구 월 추가로 한 칸 밀림
            combined_notes = "\n".join(filter(None, [extra_note, excel_note])) or None

            course_name = name_cleaned
            if not course_name: continue
            # 과정명이 순수 숫자이거나 너무 짧으면 불필요한 행으로 간주 → 건너뜀
            if course_name.isdigit() or len(course_name) < 2: continue

            # 촬영날짜에서 월 자동 추출
            shoot_date = to_date(row[13])
            auto_month = f"{shoot_date.month}월" if shoot_date else prev_month

            # 수동 입력 데이터 복원 (replace 모드에서 기존 데이터 유지)
            manual = manual_data.get(course_name.strip(), {})

            db.add(Content(year=year, shooting_month=auto_month or None,
                course_name=course_name, required_optional=str(row[3] or "") or None,
                original_code=str(row[4] or "") or None, category=str(row[5] or "") or None,
                course_code=str(row[6] or "") or None,
                session_count=int(row[7]) if row[7] and str(row[7]).isdigit() else None,
                chapter_count=int(row[8]) if row[8] and str(row[8]).isdigit() else None,
                instructor=str(row[9] or "") or None, department=str(row[10] or "") or None,
                kicpa_manager=str(row[11] or "") or None, filming_consent=str(row[12] or "") or None,
                shooting_date=shoot_date, shooting_time=str(row[14] or "") or None,
                shooting_format=str(row[15] or "") or None, location=str(row[16] or "") or None,
                has_quiz=str(row[17] or "") or None,
                quiz_count=int(row[18]) if row[18] and str(row[18]).isdigit() else None,
                materials_supply=str(row[19] or "") or None, video_marking=str(row[20] or "") or None,
                dev_outsource_date=to_date(row[21]), inspection_date=to_date(row[22]),
                open_date=to_date(row[23]), billing=str(row[24] or "") or None,
                billing_month=normalize_month(str(row[25] or "")),
                # 수동 입력값: Excel에 없는 항목이므로 기존 값 우선 복원
                custom_price=manual.get("custom_price"),
                travel_hours=manual.get("travel_hours"),
                travel_expense=manual.get("travel_expense"),
                # 비고: Excel 값 우선, 없으면 기존 수동 입력 비고 유지
                notes=combined_notes or manual.get("notes")))
            imported += 1

        db.commit()
        msg = f"{imported}건의 콘텐츠를 가져왔습니다."
        msg_type = "success"
    except Exception as e:
        msg = f"오류: {str(e)}"
        msg_type = "danger"

    return templates.TemplateResponse("import.html", {"request": request, "user": get_user(request), "msg": msg, "msg_type": msg_type})

# ── 정산 관리 (관리자) ────────────────────────────
@app.get("/billing", response_class=HTMLResponse)
def billing_page(request: Request, year: int = 2026, month: str = "", dept: str = "", db: Session = Depends(get_db)):
    require_admin(request)
    q = db.query(Content).filter_by(year=year)
    if month: q = q.filter_by(billing_month=month)
    if dept:  q = q.filter_by(department=dept)
    from sqlalchemy import text as sa_text
    billing_month_sort = case(
        {m: i for i, m in enumerate(MONTHS)},
        value=Content.billing_month, else_=99)
    date_sort = sa_text(
        "COALESCE(shooting_date, "
        "  CASE WHEN video_marking ~ '^[0-9]{4}[./-][0-9]' "
        "  THEN TO_DATE(SUBSTRING(REPLACE(REPLACE(video_marking,'.', '-'),'/','-'),1,10),'YYYY-MM-DD') "
        "  ELSE NULL END"
        ") DESC NULLS LAST"
    )
    contents = q.order_by(billing_month_sort.desc(), date_sort, Content.id.desc()).all()

    price_tbl = {p.type_name: p.unit_price for p in db.query(PriceTable).filter_by(is_active=True).all()}

    def match_price(fmt):
        """촬영형식 문자열로 단가표 단가 조회"""
        if not fmt: return 0
        # 포팅 계열: 편집 여부로 분기
        if "포팅" in fmt:
            if "편집" in fmt:
                return price_tbl.get("편집포팅", 0) or 0
            else:
                return price_tbl.get("포팅", 0) or 0
        # 출장: FullVod (출장) 단가 적용
        if "출장" in fmt:
            return price_tbl.get("FullVod (출장)", 0) or 0
        # 그 외: 단가표 type_name이 촬영형식에 포함되는지 확인
        for k, v in price_tbl.items():
            if k in fmt: return v or 0
        return 0

    # 출장비 시간당 단가 (단가표 travel 카테고리 "1 ~ 4시간")
    travel_rate = price_tbl.get("1 ~ 4시간", 100000)

    def get_unit_price(content_row):
        """별도 단가 우선, 없으면 단가표에서 촬영형식으로 조회"""
        if content_row.custom_price:
            return content_row.custom_price
        return match_price(content_row.shooting_format or "")

    def get_travel_expense(content_row):
        """출장비 계산
        우선순위: 직접입력 > 시간계산(1일 4시간 초과 시 4시간 캡) > 기본 1시간
        """
        if not content_row.shooting_format or "출장" not in content_row.shooting_format:
            return 0
        if content_row.travel_expense is not None:
            return content_row.travel_expense
        if content_row.travel_hours:
            hours = content_row.travel_hours
            days  = getattr(content_row, 'travel_days', None)
            # 1일 4시간 초과분 캡 적용
            capped = min(hours, days * 4) if days and days > 0 else hours
            return capped * travel_rate
        return travel_rate  # 기본 1시간

    # billing 템플릿 하위 호환용
    def get_price(fmt):
        return match_price(fmt or "")

    summary = {}
    for r in contents:
        d = r.department or "미지정"
        if d not in summary: summary[d] = {"count":0,"sessions":0,"total":0,"billed":0}
        p = get_unit_price(r) * (r.session_count or 0) + get_travel_expense(r)
        summary[d]["count"] += 1
        summary[d]["sessions"] += r.session_count or 0
        summary[d]["total"] += p
        if r.billing: summary[d]["billed"] += p

    depts = [d[0] for d in db.query(Content.department).filter_by(year=year).filter(Content.department != None).distinct().order_by(Content.department).all()]
    # 비용청구 월 목록 (실제 데이터 기준)
    billing_months = [m[0] for m in db.query(Content.billing_month).filter_by(year=year).filter(Content.billing_month != None).distinct().all()]
    billing_months.sort(key=lambda m: MONTH_ORDER.get(m, 99))
    return templates.TemplateResponse("billing.html", {
        "request": request, "user": get_user(request),
        "year": year, "contents": contents, "summary": summary,
        "month": month, "dept": dept, "depts": depts,
        "billing_months": billing_months,
        "get_price": get_price, "get_unit_price": get_unit_price,
        "get_travel_expense": get_travel_expense,
    })

# ── 단가표 관리 (관리자) ──────────────────────────
@app.get("/price_table", response_class=HTMLResponse)
def price_table_page(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    prices = db.query(PriceTable).order_by(PriceTable.category, PriceTable.id).all()
    return templates.TemplateResponse("price_table.html", {"request": request, "user": get_user(request), "prices": prices, "msg": ""})

@app.post("/price_table/update")
async def price_table_update(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    # 기존 항목 단가/활성여부 업데이트
    for key, val in form.items():
        if key.startswith("price_"):
            pid = int(key.split("_", 1)[1])
            p = db.query(PriceTable).filter_by(id=pid).first()
            if p:
                try: p.unit_price = int(val) if val else None
                except: p.unit_price = None
                p.is_active = f"active_{pid}" in form
    db.commit()
    return RedirectResponse("/price_table?msg=saved", 302)

@app.post("/price_table/add")
async def price_table_add(request: Request,
    category: str = Form(""), type_name: str = Form(""),
    unit_price: str = Form(""), unit: str = Form(""), note: str = Form(""),
    db: Session = Depends(get_db)):
    require_admin(request)
    if category and type_name:
        db.add(PriceTable(
            category=category, type_name=type_name,
            unit_price=int(unit_price) if unit_price else None,
            unit=unit or None, note=note or None, is_active=True))
        db.commit()
    return RedirectResponse("/price_table?msg=added", 302)

@app.post("/price_table/delete/{pid}")
def price_table_delete(request: Request, pid: int, db: Session = Depends(get_db)):
    require_admin(request)
    db.query(PriceTable).filter_by(id=pid).delete()
    db.commit()
    return RedirectResponse("/price_table?msg=deleted", 302)

# ── 사용자 관리 (관리자) ──────────────────────────
@app.get("/users", response_class=HTMLResponse)
def users_page(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    users = db.query(User).order_by(User.role, User.id).all()
    return templates.TemplateResponse("users.html", {"request": request, "user": get_user(request), "users": users, "msg": ""})

@app.post("/users/add")
def users_add(request: Request, username: str=Form(...), name: str=Form(...),
              password: str=Form(...), role: str=Form("director"), db: Session = Depends(get_db)):
    require_admin(request)
    try:
        db.add(User(username=username, password=pwd_ctx.hash(password), name=name, role=role))
        db.commit()
    except: pass
    return RedirectResponse("/users", 302)

@app.post("/users/toggle")
def users_toggle(request: Request, user_id: int=Form(...), db: Session = Depends(get_db)):
    require_admin(request)
    u = db.query(User).filter_by(id=user_id).first()
    current_user = get_user(request)
    if u and u.id != current_user["id"]:
        u.is_active = not u.is_active
        db.commit()
    return RedirectResponse("/users", 302)

@app.post("/users/change_pw")
def users_change_pw(request: Request, user_id: int=Form(...), new_password: str=Form(...), db: Session = Depends(get_db)):
    require_admin(request)
    u = db.query(User).filter_by(id=user_id).first()
    if u:
        u.password = pwd_ctx.hash(new_password)
        db.commit()
    return RedirectResponse("/users", 302)

# ── 고객담당자 관리 ───────────────────────────────
@app.get("/customers", response_class=HTMLResponse)
def customers_page(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    contacts = db.query(CustomerContact).order_by(CustomerContact.department).all()
    depts = [d[0] for d in db.query(Content.department).filter(Content.department != None).distinct().order_by(Content.department).all()]
    return templates.TemplateResponse("customers.html", {
        "request": request, "user": get_user(request),
        "contacts": contacts, "depts": depts, "msg": "",
    })

@app.post("/customers/add")
def customers_add(request: Request,
    department: str = Form(""), contact_name: str = Form(""),
    phone: str = Form(""), email: str = Form(""), note: str = Form(""),
    db: Session = Depends(get_db)):
    require_admin(request)
    if department:
        db.add(CustomerContact(department=department, contact_name=contact_name or None,
            phone=phone or None, email=email or None, note=note or None))
        db.commit()
    return RedirectResponse("/customers", 302)

@app.post("/customers/edit/{cid}")
def customers_edit(request: Request, cid: int,
    department: str = Form(""), contact_name: str = Form(""),
    phone: str = Form(""), email: str = Form(""), note: str = Form(""),
    db: Session = Depends(get_db)):
    require_admin(request)
    c = db.query(CustomerContact).filter_by(id=cid).first()
    if c:
        c.department = department or c.department
        c.contact_name = contact_name or None
        c.phone = phone or None
        c.email = email or None
        c.note = note or None
        db.commit()
    return RedirectResponse("/customers", 302)

@app.post("/customers/delete/{cid}")
def customers_delete(request: Request, cid: int, db: Session = Depends(get_db)):
    require_admin(request)
    db.query(CustomerContact).filter_by(id=cid).delete()
    db.commit()
    return RedirectResponse("/customers", 302)

# ── 스튜디오 대관료 관리 ──────────────────────────
@app.get("/studio", response_class=HTMLResponse)
def studio_page(request: Request, year: int = 2026, month: str = "", db: Session = Depends(get_db)):
    require_admin(request)
    q = db.query(StudioRental).filter_by(year=year)
    if month: q = q.filter_by(month=month)
    rentals = q.order_by(StudioRental.usage_date).all()
    months = [m[0] for m in db.query(StudioRental.month).filter_by(year=year)
              .filter(StudioRental.month != None).distinct().all()]
    months.sort(key=lambda m: MONTH_ORDER.get(m, 99))
    # 월별 소계
    monthly = {}
    for r in db.query(StudioRental).filter_by(year=year).all():
        key = r.month or "-"
        monthly.setdefault(key, {"hours": 0, "amount": 0})
        monthly[key]["hours"]  += r.hours or 0
        monthly[key]["amount"] += (r.hours or 0) * (r.unit_price or 45000)
    return templates.TemplateResponse("studio.html", {
        "request": request, "user": get_user(request),
        "year": year, "month": month, "rentals": rentals,
        "months": months, "MONTHS": MONTHS, "monthly": monthly,
    })

@app.post("/studio/add")
def studio_add(request: Request, year: int = Form(2026), month: str = Form(""),
    usage_date: str = Form(""), hours: str = Form(""),
    unit_price: str = Form("45000"), notes: str = Form(""),
    db: Session = Depends(get_db)):
    require_admin(request)
    try:
        ud = date.fromisoformat(usage_date) if usage_date else None
        m  = normalize_month(month) or (f"{ud.month}월" if ud else None)
        if ud and hours:
            db.add(StudioRental(year=year, month=m, usage_date=ud,
                hours=int(hours), unit_price=int(unit_price) if unit_price else 45000,
                notes=notes or None))
            db.commit()
    except Exception: pass
    return RedirectResponse(f"/studio?year={year}&month={m or ''}", 302)

@app.post("/studio/delete/{rid}")
def studio_delete(request: Request, rid: int, year: int = Form(2026), month: str = Form(""),
    db: Session = Depends(get_db)):
    require_admin(request)
    r = db.query(StudioRental).filter_by(id=rid).first()
    m = r.month if r else month
    db.query(StudioRental).filter_by(id=rid).delete()
    db.commit()
    return RedirectResponse(f"/studio?year={year}&month={m or ''}", 302)

# ── 문서 생성 (관리자) ────────────────────────────
@app.get("/documents", response_class=HTMLResponse)
def documents_page(request: Request, year: int = 2026, dept: str = "", month: str = "",
                   db: Session = Depends(get_db)):
    require_admin(request)
    depts = [d[0] for d in db.query(Content.department).filter_by(year=year)
             .filter(Content.department != None).distinct().order_by(Content.department).all()]
    billing_months = [m[0] for m in db.query(Content.billing_month).filter_by(year=year)
                      .filter(Content.billing_month != None).distinct().all()]
    billing_months.sort(key=lambda m: MONTH_ORDER.get(m, 99))
    # 선택된 부서+월의 과정 미리보기
    preview = []
    if dept and month:
        preview = db.query(Content).filter_by(year=year, department=dept, billing_month=month).all()
    return templates.TemplateResponse("documents.html", {
        "request": request, "user": get_user(request),
        "year": year, "depts": depts, "dept": dept,
        "billing_months": billing_months, "month": month,
        "preview": preview, "MONTHS": MONTHS,
    })

@app.post("/documents/generate")
def documents_generate(request: Request, year: int = Form(2026),
                       dept: str = Form(""), month: str = Form(""),
                       db: Session = Depends(get_db)):
    import traceback
    require_admin(request)
    if not dept or not month:
        return RedirectResponse(f"/documents?year={year}", 302)

    try:
        # 과정 목록
        courses = db.query(Content).filter_by(year=year, department=dept,
                                              billing_month=month).all()
        if not courses:
            return RedirectResponse(
                f"/documents?year={year}&dept={dept}&month={month}&msg=no_data", 302)

        # 단가표
        price_tbl = {p.type_name: p.unit_price
                     for p in db.query(PriceTable).filter_by(is_active=True).all()}

        # 스튜디오 대관료 (해당 월)
        studio_rentals = db.query(StudioRental).filter_by(year=year, month=month).all()
        studio_hours   = sum(r.hours or 0 for r in studio_rentals)

        # 이 부서가 해당 월 최대 개발비 부서인지 판별
        tr = price_tbl.get("1 ~ 4시간", 100000)
        def dept_revenue(d):
            rows = db.query(Content).filter_by(year=year, billing_month=month, department=d).all()
            return sum(
                docgen.get_unit_price_for(c, price_tbl) * (c.session_count or c.chapter_count or 0)
                + docgen.get_travel_for(c, tr)
                for c in rows
            )
        all_depts = [d[0] for d in db.query(Content.department)
                     .filter_by(year=year, billing_month=month)
                     .filter(Content.department != None).distinct().all()]
        max_dept       = max(all_depts, key=dept_revenue) if all_depts else dept
        include_studio = (dept == max_dept)

        # 고객담당자
        customer = db.query(CustomerContact).filter_by(
            department=dept, is_active=True).first()

        # ZIP 생성
        zip_bytes = docgen.generate_all(
            courses=courses, dept=dept, month_str=month, year=year,
            price_tbl=price_tbl, studio_hours=studio_hours,
            include_studio=include_studio, customer_contact=customer)

    except Exception as e:
        err_detail = traceback.format_exc()
        return templates.TemplateResponse("error.html", {
            "request": request,
            "user": get_user(request),
            "title": "문서 생성 오류",
            "message": str(e),
            "detail": err_detail,
        }, status_code=500)

    from urllib.parse import quote
    month_num  = docgen.get_month_number(month)
    dept_short = dept.replace(" ", "")
    filename   = f"{year}{month_num:02d}_한공회_{dept_short}_정산문서.zip"
    encoded    = quote(filename, safe="")
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )
